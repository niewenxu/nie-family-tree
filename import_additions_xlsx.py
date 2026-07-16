import json
import re
import argparse
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).parent
DATA_PATH = ROOT / 'data.json'

parser = argparse.ArgumentParser(description='将新增族人 Excel 合并到 data.json')
parser.add_argument('source', type=Path, help='待导入的 Excel 文件路径')
parser.add_argument('--dry-run', action='store_true', help='只显示导入结果，不写入 data.json')
args = parser.parse_args()
SOURCE = args.source.expanduser().resolve()

if not SOURCE.is_file():
    parser.error(f'找不到 Excel 文件：{SOURCE}')

data = json.loads(DATA_PATH.read_text(encoding='utf-8'))
rows = list(load_workbook(SOURCE, read_only=True, data_only=True).active.iter_rows(min_row=2, values_only=True))

def normalize(value):
    return re.sub(r'\s+', '', str(value or '').replace('聂', '聶'))

existing = {normalize(person['n']) for generation in data['zupu'] for person in generation['m']}
pending_existing = {normalize(person['n']) for person in data.get('pending', [])}

known_special = {
    '聶爱忠': 12,
    '聶爱华': 12,
    '聶伏雨': 13,
    '聶东升': 14,
    '聶磊': 14,
}
alias_only = {'聶文旭'}
generation_prefixes = {'德': 11, '广': 12, '成': 13, '先': 14, '印': 15}

added_generations = []
added_pending = []
skipped_existing = []
skipped_duplicate_input = []
seen = Counter()

for raw_name, raw_birth in rows:
    name = normalize(raw_name)
    birth = str(raw_birth).strip() if raw_birth else None
    seen[name] += 1
    if seen[name] > 1:
        skipped_duplicate_input.append((raw_name, birth))
        continue
    if name in alias_only or name in existing or name in pending_existing:
        skipped_existing.append((raw_name, birth))
        continue

    generation_index = known_special.get(name)
    if generation_index is None and len(name) >= 2:
        generation_index = generation_prefixes.get(name[1])

    record = {'n': name, 's': ''}
    if birth:
        record['birth'] = birth

    if generation_index is not None and 1 <= generation_index <= len(data['zupu']):
        # generation_index is a human-facing, one-based generation number.
        generation = data['zupu'][generation_index - 1]
        generation['m'].append(record)
        existing.add(name)
        added_generations.append((raw_name, generation['g'], birth))
    else:
        data.setdefault('pending', []).append(record)
        pending_existing.add(name)
        added_pending.append((raw_name, birth))

# From the 广 generation (12th generation) onward: undated records first in
# stable order, then dated records from oldest to youngest.
for generation_index in range(11, len(data['zupu'])):
    people = data['zupu'][generation_index]['m']
    unknown = [person for person in people if not person.get('birth')]
    dated = sorted((person for person in people if person.get('birth')), key=lambda person: person['birth'])
    data['zupu'][generation_index]['m'] = unknown + dated

pending_people = data.get('pending', [])
pending_dated = sorted((person for person in pending_people if person.get('birth')), key=lambda person: person['birth'])
pending_unknown = [person for person in pending_people if not person.get('birth')]
data['pending'] = pending_dated + pending_unknown

if not args.dry_run:
    DATA_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
print(json.dumps({
    'dry_run': args.dry_run,
    'added_generations': added_generations,
    'added_pending': added_pending,
    'skipped_existing_count': len(skipped_existing),
    'skipped_duplicate_input': skipped_duplicate_input,
    'total_people': sum(len(g['m']) for g in data['zupu']),
    'pending_count': len(data.get('pending', [])),
}, ensure_ascii=False, indent=2))
