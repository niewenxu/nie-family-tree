import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).parent
DATA = json.loads((ROOT / "data.json").read_text(encoding="utf-8"))
OUT = Path(sys.argv[1]).expanduser() if len(sys.argv) > 1 else ROOT / "聂氏宗谱维护数据库_2026版.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "族人数据库"

headers = [
    "序号", "世系数字", "世系", "字辈", "姓", "名", "姓名完整", "字", "号",
    "出生年月日", "卒年月日", "配偶", "子女", "联系方式", "迁徙地", "功名事迹", "墓葬",
    "简介", "资料来源", "备注"
]
ws.append(headers)

serial = 1
for generation_index, generation in enumerate(DATA["zupu"], start=1):
    generation_name = generation["g"]
    generation_char = DATA["beifen"][generation_index - 1].get("char") or "待考证"
    for person in generation["m"]:
        full_name = str(person.get("n") or "").replace("聶", "聂").replace("璽", "玺").replace("廣", "广").replace("潤", "润").replace(" ", "")
        given_name = full_name[1:] if full_name.startswith("聂") else full_name
        spouse = person.get("s")
        if not spouse or spouse == "未知":
            spouse = "待考证"
        death = person.get("d") or "待考证"
        bio = person.get("bio") or person.get("i") or "待考证"
        ws.append([
            serial, generation_index, generation_name, generation_char, "聂", given_name, full_name,
            person.get("zi") or "待考证", person.get("hao") or "待考证",
            person.get("birth") or "待考证", death, spouse,
            person.get("children") or "待考证", person.get("contact") or "待考证", person.get("migration") or "待考证",
            person.get("achievements") or bio, person.get("burial") or "待考证",
            bio, person.get("source") or "待考证", person.get("note") or "待考证"
        ])
        serial += 1

for person in DATA.get("pending", []):
    full_name = str(person.get("n") or "").replace("聶", "聂").replace(" ", "")
    given_name = full_name[1:] if full_name.startswith("聂") else full_name
    spouse = person.get("s")
    if not spouse or spouse == "未知":
        spouse = "待考证"
    ws.append([
        serial, "", "世系待考", "待考证", "聂", given_name, full_name,
        person.get("zi") or "待考证", person.get("hao") or "待考证",
        person.get("birth") or "待考证", person.get("d") or "待考证", spouse,
        person.get("children") or "待考证", person.get("contact") or "待考证", person.get("migration") or "待考证",
        person.get("achievements") or "待考证", person.get("burial") or "待考证",
        person.get("bio") or person.get("i") or "待考证",
        person.get("source") or "待考证", person.get("note") or "待考证"
    ])
    serial += 1

# Reserve styled/formula-ready rows. Excel tables propagate formulas and styling when users add records.
for row in range(ws.max_row + 1, ws.max_row + 101):
    ws.cell(row, 1, f'=IF(F{row}="","",ROW()-1)')
    for col in range(2, len(headers) + 1):
        ws.cell(row, col, "")
    ws.cell(row, 7, f'=IF(F{row}="","",E{row}&F{row})')

table = Table(displayName="FamilyDatabase", ref=f"A1:T{ws.max_row}")
table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
ws.add_table(table)

header_fill = PatternFill("solid", fgColor="263B32")
header_font = Font(name="Songti SC", color="F6F1E7", bold=True, size=11)
thin = Side(style="thin", color="D8CCB3")
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(bottom=thin)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
    for cell in row:
        cell.font = Font(name="Songti SC", size=10, color="25251F")
        cell.alignment = Alignment(vertical="top", wrap_text=True)

widths = {"A": 9, "B": 10, "C": 10, "D": 10, "E": 7, "F": 14, "G": 16, "H": 12, "I": 12, "J": 15, "K": 15, "L": 18, "M": 22, "N": 18, "O": 20, "P": 34, "Q": 24, "R": 38, "S": 24, "T": 24}
for col, width in widths.items():
    ws.column_dimensions[col].width = width
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:T{ws.max_row}"
ws.row_dimensions[1].height = 28

gen_validation = DataValidation(type="list", formula1='"始祖,二世,三世,四世,五世,六世,七世,八世,九世,十世,十一世,十二世,十三世,十四世,十五世,十六世,十七世,十八世,十九世,二十世"', allow_blank=True)
ws.add_data_validation(gen_validation)
gen_validation.add(f"C2:C{ws.max_row}")

guide = wb.create_sheet("填写说明")
guide_rows = [
    ("聂氏宗谱维护数据库", "用于维护族人姓名、世系与人物资料"),
    ("新增族人", "在“族人数据库”表格末行继续输入；序号会随表格扩展自动填充。"),
    ("世系数字", "填写阿拉伯数字 1—20；页面按此字段自动归入对应世系。"),
    ("缺失信息", "统一填写“待考证”，不要留空或使用“未知”“暂无”。"),
    ("联系方式", "仅供族人详情使用；普通号码前端始终隐藏后4位，聂强在族谱密码解锁后可完整查看。"),
    ("姓名规则", "姓、名分列填写；名中无需重复填写“聂”。"),
    ("日期规则", "能确认具体日期时使用 YYYY-MM-DD；仅知年份可填写年份并在备注说明。"),
    ("人物字段", "字辈、名、字、号、生卒年月日、配偶、子女、迁徙地、功名事迹、墓葬、简介均可直接维护。"),
    ("自动更新", "页面导入本表后，将按现有行数重新计算族人数量，并按世系数字排序和显示。")
]
for row in guide_rows:
    guide.append(row)
guide.column_dimensions["A"].width = 22
guide.column_dimensions["B"].width = 88
for cell in guide[1]:
    cell.fill = header_fill
    cell.font = Font(name="STKaiti", color="F6F1E7", bold=True, size=14)
for row in guide.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(bottom=thin)
guide.freeze_panes = "A2"

wb.save(OUT)

# Re-open as a structural check.
check = load_workbook(OUT, data_only=False)
assert "族人数据库" in check.sheetnames and check["族人数据库"].max_column == len(headers)
assert check["族人数据库"].max_row == len([p for g in DATA["zupu"] for p in g["m"]]) + len(DATA.get("pending", [])) + 101
print(OUT)
