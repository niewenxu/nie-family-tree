import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).parent
WORKBOOK = ROOT / '聂氏宗谱数据库_2026-07-12_含完整姓名.xlsx'
DATA_PATH = ROOT / 'data.json'

def normalize(value):
    return re.sub(r'\s+', '', str(value or '').replace('聂', '聶'))

data = json.loads(DATA_PATH.read_text(encoding='utf-8'))
people = [person for generation in data['zupu'] for person in generation['m']] + data.get('pending', [])
index = {}
for person in people:
    index.setdefault(normalize(person['n']), []).append(person)

ws = load_workbook(WORKBOOK, read_only=True, data_only=True)['族人数据库']
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
synced = []
unmatched = []
for values in ws.iter_rows(min_row=2, values_only=True):
    row = dict(zip(headers, values))
    full_name = row.get('姓名完整')
    contact = row.get('联系方式')
    if not full_name or not contact:
        continue
    matches = index.get(normalize(full_name), [])
    if len(matches) != 1:
        unmatched.append({'姓名': full_name, '匹配数': len(matches)})
        continue
    contact_text = re.sub(r'\.0$', '', str(contact).strip())
    matches[0]['contact'] = contact_text
    synced.append(full_name)

DATA_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
print(json.dumps({'synced': len(synced), 'unmatched': unmatched}, ensure_ascii=False, indent=2))
